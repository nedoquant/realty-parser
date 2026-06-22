# -*- coding: utf-8 -*-
"""Единый лист «Динамика цен»: цены CIAN + Avito с историей по месяцам.
Формат — ДАТЫ ВПРАВО: слева фикс. ключ (Источник|Город|Категория|Тип цены),
каждый месяц = ОДИН столбец справа, последний столбец Δ% = изменение к прошлому
месяцу (формула). Новый прогон месяца добавляет/обновляет свой столбец.

Использование из парсеров:
    import price_dynamics as pd
    pd.update_dynamics("cian_объявления.xlsx", "Avito", "2026-06-16", records)
где records = [(категория, город, тип_цены, значение), ...] (категория="" если нет).
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import config

OUT = config.OUTPUT_FILE

SHEET = "Динамика цен"
KEYS = ["Источник", "Город", "Категория", "Тип цены"]   # 4 ключевых столбца слева
NK = len(KEYS)
MONTHS_RU = ["янв", "фев", "мар", "апр", "май", "июн",
             "июл", "авг", "сен", "окт", "ноя", "дек"]
HEAD_FILL = PatternFill("solid", fgColor="DDE7F0")


def month_label(date_iso):
    """'2026-06-16' -> 'июн.2026' (группировка по месяцу)."""
    y, m = int(date_iso[:4]), int(date_iso[5:7])
    return f"{MONTHS_RU[m - 1]}.{y}"


def _date_cols(ws):
    """{label: col} для всех столбцов-месяцев (заголовок не из KEYS и не 'Δ%')."""
    out = {}
    for c in range(NK + 1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h and h != "Δ%":
            out[h] = c
    return out


def _ensure_sheet(wb):
    if SHEET in wb.sheetnames:
        return wb[SHEET]
    ws = wb.create_sheet(SHEET)
    for i, h in enumerate(KEYS, 1):
        ws.cell(1, i, h)
    return ws


def _row_index(ws):
    """Ключ строки с нормализацией None->'' (пустая «Категория» хранится как None,
    но в ключе сравнивается как '')."""
    idx = {}
    for r in range(2, ws.max_row + 1):
        raw = [ws.cell(r, i).value for i in range(1, NK + 1)]
        if any(v is not None for v in raw):
            idx[tuple(v if v is not None else "" for v in raw)] = r
    return idx


def _style(ws):
    """Шапка, заморозка, автофильтр, ширины, цвет Δ% (зелёный рост/красный спад)."""
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = {1: 10, 2: 18, 3: 34, 4: 30}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for c in range(NK + 1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = "E2"                     # видны ключ-столбцы и шапка
    last = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last}{ws.max_row}"
    # цвет Δ% (последний столбец)
    dl = get_column_letter(ws.max_column)
    rng = f"{dl}2:{dl}{ws.max_row}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"], font=Font(color="0A7A0A")))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"], font=Font(color="C00000")))


def update_dynamics(path, source, date_iso, records):
    """Добавить/обновить столбец месяца в листе «Динамика цен». records =
    [(категория, город, тип_цены, значение)]. Возвращает имя столбца месяца."""
    wb = openpyxl.load_workbook(path)
    ws = _ensure_sheet(wb)
    lbl = month_label(date_iso)
    dcols = _date_cols(ws)

    # столбец месяца: существующий (повторный прогон) или новый справа от месяцев
    if lbl in dcols:
        tcol = dcols[lbl]
    else:
        tcol = (max(dcols.values()) if dcols else NK) + 1
        # этот столбец мог занимать старый Δ% — чистим его целиком, ставим заголовок
        for r in range(1, ws.max_row + 1):
            ws.cell(r, tcol).value = None
        ws.cell(1, tcol, lbl)
        dcols[lbl] = tcol

    idx = _row_index(ws)
    append_at = ws.max_row + 1
    for cat, city, typ, val in records:
        key = (source, city, cat or "", typ)
        r = idx.get(key)
        if r is None:
            r = append_at
            append_at += 1
            for i, v in enumerate(key, 1):
                ws.cell(r, i, v)
            idx[key] = r
        ws.cell(r, tcol).value = val

    # пересобрать Δ% в самом правом столбце (после всех месяцев)
    dcols = _date_cols(ws)
    ordered = sorted(dcols.values())                 # хронологически слева-направо
    # убрать старые столбцы Δ%
    for c in range(NK + 1, ws.max_column + 2):
        if ws.cell(1, c).value == "Δ%":
            for r in range(1, ws.max_row + 1):
                ws.cell(r, c).value = None
    delta_col = ordered[-1] + 1
    ws.cell(1, delta_col, "Δ%")
    if len(ordered) >= 2:
        last_l = get_column_letter(ordered[-1])
        prev_l = get_column_letter(ordered[-2])
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, ordered[-1]).value is None and ws.cell(r, ordered[-2]).value is None:
                continue
            f = (f'=IFERROR(({last_l}{r}-{prev_l}{r})/{prev_l}{r},"")')
            cell = ws.cell(r, delta_col, f)
            cell.number_format = "0.0%"

    _style(ws)
    wb.save(path)
    return lbl


# ---------- разовый перенос текущих листов CIAN Price / Avito Price ----------
def _avito_records(wb):
    ws = wb["Avito Price"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    date = None
    recs = []
    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, 1).value
        if not city:
            continue
        for c in range(2, ws.max_column):           # без 'Обновлено'
            h = headers[c - 1]
            if h == "Обновлено":
                continue
            v = ws.cell(r, c).value
            if v is not None:
                recs.append(("", city, h, v))
        date = date or ws.cell(r, ws.max_column).value
    return date or "2026-06-16", recs


def _cian_records(wb):
    ws = wb["CIAN Price"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # 1 Сделка 2 Тип 3 Категория 4 Город 5..8 цены 9 Примечание 10 Обновлено
    price_cols = [c for c in range(5, 9)]
    date = None
    recs = []
    for r in range(2, ws.max_row + 1):
        deal, ptype, cat, city = (ws.cell(r, i).value for i in range(1, 5))
        if not city:
            continue
        catlabel = " · ".join(str(x) for x in (deal, ptype, cat) if x)
        for c in price_cols:
            v = ws.cell(r, c).value
            if v is not None:
                recs.append((catlabel, city, headers[c - 1], v))
        date = date or ws.cell(r, 10).value
    return date or "2026-06-10", recs


def sync_avito(path=OUT):
    """Дописать свежие цены Avito (из листа «Avito Price») в «Динамика цен»."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Avito Price" not in wb.sheetnames:
        return None
    date, recs = _avito_records(wb)
    n = update_dynamics(path, "Avito", date, recs)
    print(f"Динамика цен: Avito {len(recs)} цен -> {n}")
    return n


def sync_cian(path=OUT):
    """Дописать свежие цены CIAN (из листа «CIAN Price») в «Динамика цен»."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "CIAN Price" not in wb.sheetnames:
        return None
    date, recs = _cian_records(wb)
    n = update_dynamics(path, "CIAN", date, recs)
    print(f"Динамика цен: CIAN {len(recs)} цен -> {n}")
    return n


def migrate(path=OUT):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    if "Avito Price" in wb.sheetnames:
        out["Avito"] = _avito_records(wb)
    if "CIAN Price" in wb.sheetnames:
        out["CIAN"] = _cian_records(wb)
    for source, (date, recs) in out.items():
        n = update_dynamics(path, source, date, recs)
        print(f"{source}: {len(recs)} цен -> столбец {n}")


if __name__ == "__main__":
    migrate()
