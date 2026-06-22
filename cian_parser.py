# -*- coding: utf-8 -*-
"""
Парсер количества объявлений с ЦИАН по матрице КАТЕГОРИЯ × ГОРОД.

Для каждой категории и каждого города берёт ОДНО число — сколько найдено
объявлений ("Найдено N объявлений"). Результат дописывается в Excel:
строки = (Группа, Категория, Город), столбцы = даты запусков (история копится).

Запуск:  python cian_parser.py   (или двойным кликом по run.bat)

Почему curl: ЦИАН блокирует обычные Python-запросы по TLS-отпечатку (кидает
на капчу), а системный curl проходит свободно.
"""

import re
import os
import sys
import time
import datetime
import subprocess

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn

import config

# вывод в UTF-8 (иначе кириллица/× ломаются при перенаправлении в файл на Windows)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:  # noqa: BLE001
    pass

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36")

# ----------------------------------------------------------------------------
# Загрузка страницы и извлечение числа
# ----------------------------------------------------------------------------

def fetch_html(url, proxy=None, cookiejar=None, referer=None):
    """Скачивает страницу через curl. Возвращает (html, final_url).
    -g отключает globbing (квадратные скобки в URL). --compressed просит gzip.
    proxy — строка для curl -x. cookiejar — файл cookie (-c/-b): держит СЕССИЮ
    между запросами (как браузер) — мягче к анти-боту, чем запросы «без памяти».
    referer — заголовок Referer (Авито пускает только запросы «со ссылкой»)."""
    cmd = ["curl", "-sgL", "--compressed",
           "-A", UA,
           "-H", "Accept-Language: ru-RU,ru;q=0.9,en;q=0.8",
           "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,"
                 "image/avif,image/webp,*/*;q=0.8",
           "-w", "\n@@@%{url_effective}"]
    if referer:
        cmd += ["-H", f"Referer: {referer}"]
    if cookiejar:
        cmd += ["-c", cookiejar, "-b", cookiejar]
    if proxy:
        cmd += ["-x", proxy]
    cmd.append(url)
    result = subprocess.run(cmd, capture_output=True, timeout=45)
    text = result.stdout.decode("utf-8", "replace")
    body, _, final_url = text.rpartition("\n@@@")
    return body, final_url


def extract_count(html):
    """Достаёт количество объявлений. Возвращает int или None."""
    # 1) Основной способ — число в JSON страницы
    m = re.search(r'"totalOffers":(\d+)', html)
    if m:
        return int(m.group(1))
    # 2) Запасной способ — из видимого текста "Найдено N ..." (спец-объекты).
    #    Число может содержать пробелы/неразрывные пробелы: "27 209".
    m = re.search(r"Найдено\s*([\d   \.]+?)\s*"
                  r"(?:объявлен|предложен|вариант)", html)
    if m:
        digits = re.sub(r"\D", "", m.group(1))
        if digits:
            return int(digits)
    return None


def parse_one(url):
    """Парсит один URL. Возвращает (count_or_None, status_text)."""
    try:
        html, final_url = fetch_html(url)
    except subprocess.TimeoutExpired:
        return None, "таймаут"
    except Exception as e:  # noqa: BLE001
        return None, f"ошибка: {e}"
    if "cian-captcha" in final_url:
        return None, "КАПЧА"
    count = extract_count(html)
    if count is None:
        return None, "число не найдено"
    return count, "ок"


# ----------------------------------------------------------------------------
# Сбор данных по всей России (через JSON-API api.cian.ru)
# ----------------------------------------------------------------------------

def _fetch_regions():
    """Список (id, имя) всех 85 субъектов РФ (через curl — www.cian.ru блокирует
    requests). К коротким названиям добавляем тип («область»/«край») для ясности."""
    import json
    out = subprocess.run(["curl", "-sgL", "-A", UA, config.REGIONS_URL],
                         capture_output=True, timeout=30).stdout.decode("utf-8", "replace")
    items = json.loads(out)["data"]["items"]
    return [(it["id"], it.get("fullName") or it["name"]) for it in items]


def _api_count(session, jq):
    """offerCount по фильтру jq через API (с повторами). requests работает с
    api.cian.ru напрямую (в отличие от www.cian.ru)."""
    import json
    for _ in range(4):
        try:
            r = session.post(config.SEARCH_API,
                             data=json.dumps({"jsonQuery": jq}), timeout=30)
            return r.json()["data"]["offerCount"]
        except Exception:  # noqa: BLE001
            time.sleep(1.0)
    return None


def collect_russia():
    """Считает «Вся Россия» по каждой категории. Для типов аренды (flatrent/
    suburbanrent) — сумма по 85 субъектам; для остальных хватает одного запроса
    без региона. Возвращает список (group, name, RUSSIA_LABEL, count_or_None)."""
    import requests
    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Content-Type": "application/json",
                            "Accept": "application/json"})
    regions = _fetch_regions()
    print(f"Сбор по РФ: {len(regions)} субъектов загружено.")
    print("=" * 64)

    results = []
    for group, name, *_ in config.CATEGORIES:
        jq = config.API_QUERY.get((group, name))
        if jq is None:  # спец-объекты — нет чистого фильтра, РФ пропускаем
            print(f"[{group}] {name}: пропуск (нет API-фильтра)")
            continue
        if jq["_type"] in config.RENT_SUM_TYPES:
            total, fails = 0, 0
            for rid, _ in regions:
                q = dict(jq)
                q["region"] = {"type": "terms", "value": [rid]}
                c = _api_count(session, q)
                if c is None:
                    fails += 1
                else:
                    total += c
                time.sleep(0.05)
            count = total
            note = f"  (сумма 85 рег., сбоев {fails})" if fails else "  (сумма 85 рег.)"
        else:
            count = _api_count(session, jq)
            note = "  (без региона)"
        shown = f"{count:,}".replace(",", " ") if count is not None else "—"
        print(f"[{group}] {name}: {shown}{note}")
        results.append((group, name, config.RUSSIA_LABEL, count))
    return results


def collect_regions():
    """Разбивка по всем 85 субъектам РФ для каждой категории (через API,
    по запросу на регион). Возвращает список (group, name, регион, count)."""
    import requests
    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Content-Type": "application/json",
                            "Accept": "application/json"})
    regions = _fetch_regions()
    cats = [(g, n) for g, n, *_ in config.CATEGORIES if (g, n) in config.API_QUERY]
    total_tasks = len(cats) * len(regions)
    print(f"Разбивка по РФ: {len(cats)} категорий × {len(regions)} субъектов "
          f"= {total_tasks} запросов")
    print("=" * 64)

    results, done = [], 0
    for group, name in cats:
        jq = config.API_QUERY[(group, name)]
        s, fails = 0, 0
        for rid, rname in regions:
            q = dict(jq)
            q["region"] = {"type": "terms", "value": [rid]}
            c = _api_count(session, q)
            if c is None:
                fails += 1
            else:
                s += c
            results.append((group, name, rname, c))
            done += 1
            time.sleep(0.05)
        shown = f"{s:,}".replace(",", " ")
        print(f"[{group}] {name}: Σ={shown}" + (f"  (сбоев {fails})" if fails else ""))
    return results


# ----------------------------------------------------------------------------
# Запись в Excel с накоплением истории
# ----------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GROUP_FONT = Font(bold=True, color="1F3864")
CAT_FONT = Font(bold=True)
BLOCK_FILL = PatternFill("solid", fgColor="EAF0FA")
RUSSIA_FILL = PatternFill("solid", fgColor="FCE9C8")  # строка «Вся Россия»
RUSSIA_FONT = Font(bold=True, color="7A4F12")
THIN = Side(style="thin", color="D9D9D9")
TOPSEP = Side(style="medium", color="9DB7E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def open_or_create(path, sheet_name):
    try:
        wb = openpyxl.load_workbook(path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
    except PermissionError:
        # файл занят (открыт в Excel) — начинаем с чистой книги,
        # сохранение позже уйдёт в основной или запасной файл
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb


def _read_existing(ws):
    """Считывает уже накопленные данные. Возвращает (history, date_headers),
    где history[(группа,категория,город)] = {дата: значение}."""
    history, date_headers = {}, []
    if ws["A1"].value is None:
        return history, date_headers
    cols = {}
    c = 4
    while ws.cell(1, c).value is not None:
        h = str(ws.cell(1, c).value)
        date_headers.append(h)
        cols[c] = h
        c += 1
    r = 2
    while ws.cell(r, 3).value is not None:
        key = (ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
        rec = history.setdefault(key, {})
        for col, h in cols.items():
            v = ws.cell(r, col).value
            if v is not None:
                rec[h] = v
        r += 1
    return history, date_headers


def _ordered_keys(history):
    """Канонический порядок строк: по категориям из config, внутри блока —
    сначала «Вся Россия», затем 16 городов. Лишние ключи добавляются в конец."""
    geos = [config.RUSSIA_LABEL] + [c[0] for c in config.CITIES]
    ordered, seen = [], set()
    for group, name, *_ in config.CATEGORIES:
        for geo in geos:
            k = (group, name, geo)
            if k in history:
                ordered.append(k)
                seen.add(k)
    for k in history:
        if k not in seen:
            ordered.append(k)
    return ordered


def _merge_history(ws, results, date_label):
    """Читает накопленное, домешивает новые значения. Возвращает (history, dates)."""
    history, date_headers = _read_existing(ws)
    for group, name, geo, count in results:
        if count is None:
            continue  # неудачный запрос не затирает накопленное значение
        history.setdefault((group, name, geo), {})[date_label] = count
    return history, sorted(set(date_headers) | {date_label})


def write_results(results, date_label):
    """Города. results — (group, name, city|«Вся Россия», count). Лист config.SHEET_NAME."""
    wb = open_or_create(config.OUTPUT_FILE, config.SHEET_NAME)
    ws = wb[config.SHEET_NAME]
    history, date_headers = _merge_history(ws, results, date_label)
    ordered = _ordered_keys(history)
    _save_matrix(ws, "ЦИАН_данные", "Город", history, date_headers, ordered,
                 russia_label=config.RUSSIA_LABEL, exclude_label=config.RUSSIA_LABEL)
    return _safe_save(wb, config.OUTPUT_FILE)


def write_regions(results, date_label):
    """Регионы. results — (group, name, регион, count). Отдельный лист со всеми
    субъектами РФ по категориям (для понимания распределения по стране)."""
    wb = open_or_create(config.OUTPUT_FILE, config.REGIONS_SHEET)
    ws = wb[config.REGIONS_SHEET]
    history, date_headers = _merge_history(ws, results, date_label)
    # порядок: категории из config, внутри — регионы в порядке появления в results
    region_order = []
    for _, _, geo, _ in results:
        if geo not in region_order:
            region_order.append(geo)
    ordered, seen = [], set()
    for group, name, *_ in config.CATEGORIES:
        for geo in region_order:
            k = (group, name, geo)
            if k in history and k not in seen:
                ordered.append(k)
                seen.add(k)
    for k in history:
        if k not in seen:
            ordered.append(k)
            seen.add(k)
    _save_matrix(ws, "ЦИАН_регионы", "Регион", history, date_headers, ordered,
                 russia_label=None, exclude_label=None)
    return _safe_save(wb, config.OUTPUT_FILE)


def write_prices(rows, date_label):
    """Снимок цен публикации на отдельный лист «CIAN Price» (перезаписывается).
    rows — список dict: deal, offer, category, city, vals{тариф:цена}, note."""
    wb = open_or_create(config.OUTPUT_FILE, config.PRICE_SHEET)
    ws = wb[config.PRICE_SHEET]
    for tname in list(ws.tables):
        del ws.tables[tname]
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    tariffs = ["Публикация", "Топ", "Премиум", "Стандарт"]
    units = {"Публикация": "₽/30 дн", "Топ": "₽/сут",
             "Премиум": "₽/сут", "Стандарт": "₽/сут"}
    titles = (["Сделка", "Тип недвижимости", "Категория", "Город"]
              + [f"{t}, {units[t]}" for t in tariffs]
              + ["Примечание", "Обновлено"])
    for col, title in enumerate(titles, 1):
        c = ws.cell(1, col, title)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(horizontal="center" if col >= 5 else "left",
                                wrap_text=True, vertical="center")
    ws.freeze_panes = "E2"
    for col, w in zip(range(1, len(titles) + 1),
                      [20, 18, 30, 18, 13, 11, 12, 13, 32, 12]):
        ws.column_dimensions[get_column_letter(col)].width = w

    # порядок: сделка -> тип -> категория(в порядке API) -> город(как в CITIES)
    deal_order = [l for _, l in config.PRICE_DEALS]
    offer_order = [l for _, l in config.PRICE_OFFERS]
    city_order = [n for n, _, _ in config.CITIES]
    cat_seq = {}
    for r in rows:
        cat_seq.setdefault((r["deal"], r["offer"]), [])
        if r["category"] not in cat_seq[(r["deal"], r["offer"])]:
            cat_seq[(r["deal"], r["offer"])].append(r["category"])

    def sk(r):
        return (deal_order.index(r["deal"]), offer_order.index(r["offer"]),
                cat_seq[(r["deal"], r["offer"])].index(r["category"]),
                city_order.index(r["city"]) if r["city"] in city_order else 99)

    rows_sorted = sorted(rows, key=sk)
    prev_block = None
    for i, r in enumerate(rows_sorted, start=2):
        block = (r["deal"], r["offer"])
        new_block = block != prev_block
        ws.cell(i, 1, r["deal"])
        ws.cell(i, 2, r["offer"])
        ws.cell(i, 3, r["category"])
        ws.cell(i, 4, r["city"])
        for j, t in enumerate(tariffs, start=5):
            v = r["vals"].get(t)
            cell = ws.cell(i, j, v if v is not None else "—")
            if isinstance(v, (int, float)):
                cell.number_format = "# ##0"
            cell.alignment = Alignment(horizontal="right")
        ws.cell(i, 9, r.get("note") or "")
        ws.cell(i, 10, date_label).alignment = Alignment(horizontal="center")
        for col in range(1, len(titles) + 1):
            top = TOPSEP if new_block else THIN
            ws.cell(i, col).border = Border(left=THIN, right=THIN, top=top, bottom=THIN)
        if new_block:
            for col in (1, 2):
                ws.cell(i, col).font = CAT_FONT
                ws.cell(i, col).fill = BLOCK_FILL
            prev_block = block

    last_row = 1 + len(rows_sorted)
    last_letter = get_column_letter(len(titles))
    tab = Table(displayName="CIAN_Price", ref=f"A1:{last_letter}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=False,
        showColumnStripes=False, showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)
    return _safe_save(wb, config.OUTPUT_FILE)


def write_yandex(results, date_label):
    """Яндекс. results — (group, name, city, count). Лист config.YANDEX_SHEET,
    та же модель с историей, что и у счётчиков ЦИАН."""
    wb = open_or_create(config.OUTPUT_FILE, config.YANDEX_SHEET)
    ws = wb[config.YANDEX_SHEET]
    history, date_headers = _merge_history(ws, results, date_label)
    geos = [config.RUSSIA_LABEL] + [n for n, _, _ in config.CITIES]
    ordered, seen = [], set()
    for group, name, *_ in config.YANDEX_CATEGORIES:
        for geo in geos:
            k = (group, name, geo)
            if k in history and k not in seen:
                ordered.append(k)
                seen.add(k)
    for k in history:
        if k not in seen:
            ordered.append(k)
            seen.add(k)
    _save_matrix(ws, "Яндекс_данные", "Город", history, date_headers, ordered,
                 russia_label=config.RUSSIA_LABEL, exclude_label=config.RUSSIA_LABEL)
    return _safe_save(wb, config.OUTPUT_FILE)


def write_yandex_regions(results, date_label):
    """Разбивка Яндекса по 85 субъектам РФ — отдельный лист."""
    wb = open_or_create(config.OUTPUT_FILE, config.YANDEX_REGIONS_SHEET)
    ws = wb[config.YANDEX_REGIONS_SHEET]
    history, date_headers = _merge_history(ws, results, date_label)
    region_order = list(config.YANDEX_REGION_SLUG.keys())
    ordered, seen = [], set()
    for group, name, *_ in config.YANDEX_CATEGORIES:
        for rname in region_order:
            k = (group, name, rname)
            if k in history and k not in seen:
                ordered.append(k)
                seen.add(k)
    for k in history:
        if k not in seen:
            ordered.append(k)
            seen.add(k)
    _save_matrix(ws, "Яндекс_регионы", "Регион", history, date_headers, ordered,
                 russia_label=None, exclude_label=None)
    return _safe_save(wb, config.OUTPUT_FILE)


def _save_matrix(ws, table_name, geo_title, history, date_headers, ordered,
                 russia_label, exclude_label):
    """Полностью перезаписывает лист: шапка, строки в заданном порядке, умная
    таблица и строка «Итого»."""
    for tname in list(ws.tables):
        del ws.tables[tname]
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    titles = ["Группа", "Категория", geo_title] + date_headers
    for col, title in enumerate(titles, 1):
        cell = ws.cell(1, col, title)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER
        cell.alignment = Alignment(horizontal="center" if col >= 4 else "left")
    ws.freeze_panes = "D2"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    for col in range(4, len(titles) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    last_col = len(titles)
    for i, key in enumerate(ordered, start=2):
        group, name, geo = key
        rec = history[key]
        ws.cell(i, 1, group)
        ws.cell(i, 2, name)
        ws.cell(i, 3, geo)
        for col, h in enumerate(date_headers, start=4):
            v = rec.get(h)
            if v is None:
                continue
            cell = ws.cell(i, col, v)
            if isinstance(v, int):
                cell.number_format = "# ##0"
            cell.alignment = Alignment(horizontal="right")

    last_data_row = 1 + len(ordered)
    _style_blocks(ws, last_data_row + 1, last_col, russia_label)
    _build_table(ws, last_data_row, last_col, table_name, exclude_label)


def _build_table(ws, last_data_row, last_col, table_name="ЦИАН_данные",
                 exclude_label=None):
    """Создаёт умную таблицу (для фильтров) на области данных и строку «Итого»
    под ней. Итог суммирует ТОЛЬКО видимые (отфильтрованные) строки; если задан
    exclude_label — исключает строки с таким значением в столбце «Город/Регион»
    (чтобы «Вся Россия» не прибавлялась к сумме городов)."""
    last_letter = get_column_letter(last_col)
    ref = f"A1:{last_letter}{last_data_row}"

    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=False,
        showColumnStripes=False, showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)

    # строка итогов под таблицей (с пустой строкой-отступом)
    totals_row = last_data_row + 2
    ws.cell(totals_row, 1, "Итого (по фильтру)").font = Font(bold=True, color="1F3864")
    hint = ("← сумма видимых строк (без «Вся Россия»)" if exclude_label
            else "← сумма видимых строк")
    ws.cell(totals_row, 3, hint).font = Font(italic=True, color="808080")
    crng = f"$C$2:$C${last_data_row}"
    for c in range(4, last_col + 1):
        L = get_column_letter(c)
        if exclude_label:
            # сумма видимых строк, КРОМЕ строк «Вся Россия»
            formula = (f'=SUMPRODUCT(SUBTOTAL(109,OFFSET({L}$2,'
                       f'ROW({L}$2:{L}${last_data_row})-ROW({L}$2),0)),'
                       f'--({crng}<>"{exclude_label}"))')
        else:
            formula = f"=SUBTOTAL(109,{L}2:{L}{last_data_row})"
        cell = ws.cell(totals_row, c, formula)
        cell.number_format = "# ##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")


def _safe_save(wb, path):
    """Сохраняет книгу. Если файл занят (открыт в Excel) — пишет в запасной
    файл с отметкой времени, чтобы не потерять результат прогона."""
    try:
        wb.save(path)
        return path
    except PermissionError:
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.replace(".xlsx", f"_{stamp}.xlsx")
        wb.save(alt)
        print(f"\n⚠ Файл {path} занят (открыт в Excel).")
        print(f"  Результат сохранён в запасной файл: {alt}")
        return alt


def _style_blocks(ws, end_row, last_col, russia_label=None):
    """Визуально разделяет блоки категорий, подсвечивает первую строку блока
    и выделяет строку «Вся Россия» (если russia_label задан)."""
    prev_cat = None
    for row in range(2, end_row):
        g = ws.cell(row, 1)
        n = ws.cell(row, 2)
        cat = (g.value, n.value)
        is_russia = russia_label is not None and ws.cell(row, 3).value == russia_label
        new_block = cat != prev_cat
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            top = TOPSEP if new_block else THIN
            cell.border = Border(left=THIN, right=THIN, top=top, bottom=THIN)
        if new_block:
            g.font, n.font = GROUP_FONT, CAT_FONT
            for col in range(1, 4):
                ws.cell(row, col).fill = BLOCK_FILL
            prev_cat = cat
        if is_russia:
            for col in range(1, last_col + 1):
                ws.cell(row, col).fill = RUSSIA_FILL
                if col != 2:
                    ws.cell(row, col).font = RUSSIA_FONT


# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------

def _avito_count(html):
    """Число объявлений со страницы Авито."""
    m = re.search(r'page-title/count[^>]*>([^<]+)', html)
    if m:
        d = re.sub(r"\D", "", m.group(1))
        if d:
            return int(d)
    m = re.search(r'"mainCount":(\d+)', html)
    return int(m.group(1)) if m else None


def _read_proxies():
    try:
        return [l.strip() for l in open(config.PROXY_FILE, encoding="utf-8")
                if l.strip() and not l.startswith("#")]
    except FileNotFoundError:
        return []


def collect_avito(only_cities=None, only_names=None, include_russia=True):
    """Количество объявлений Авито по городам × категориям. Канонические URL
    с суффиксом (плейн блокируется). curl --compressed, ретраи от анти-бота.
    only_cities — список городов; only_names — список названий категорий (добор).
    include_russia — добавить строку «вся Россия» (слаг /all/) по каждой категории.
    Если есть proxies.txt — ходит через прокси с ротацией (обход бана по IP)."""
    cities = [c for c in config.CITIES if not only_cities or c[0] in only_cities]
    cats = [c for c in config.AVITO_CATEGORIES if not only_names or c[1] in only_names]
    # geos: (подпись, слаг). Авито «вся Россия» = слаг /all/.
    geos = [(c[0], config.AVITO_CITY_SLUG.get(c[0])) for c in cities]
    if include_russia and (not only_cities or config.RUSSIA_LABEL in only_cities):
        geos.append((config.RUSSIA_LABEL, "all"))
    proxies = _read_proxies()
    if proxies:
        print(f"Прокси: {len(proxies)} шт (ротация).")
    total = len(geos) * len(cats)
    print(f"Авито: {len(geos)} гео × {len(cats)} категорий = {total} запросов")
    print("=" * 64)
    # cookie-сессия: держим cookie между запросами (как браузер) — мягче к анти-боту
    jar = "_avito_cookies.txt"
    try:
        os.remove(jar)
    except OSError:
        pass
    try:  # прогрев — получить сессионные cookie с главной
        fetch_html("https://www.avito.ru/", cookiejar=jar)
        time.sleep(1.5)
    except Exception:  # noqa: BLE001
        pass
    results, done, streak = [], 0, 0
    for city, slug in geos:
        if not slug:
            continue
        if streak >= 12:
            # IP забанен (12 блоков подряд) — нет смысла продолжать, остаёмся None
            results.append((city, "__ABORT__", None, None))
            break
        print(f"\n[{city}]")
        for group, name, path in cats:
            url = f"{config.AVITO_BASE}/{slug}/{path}"
            cnt, status = None, "блок"
            for attempt in range(4):
                proxy = proxies[(done + attempt) % len(proxies)] if proxies else None
                try:
                    html, _final = fetch_html(url, proxy=proxy, cookiejar=jar,
                                              referer="https://www.avito.ru/")
                    cnt = _avito_count(html)
                    if cnt is not None:
                        status = "ок"
                        break
                    time.sleep(2)  # анти-бот firewall — подождать и повторить
                except Exception as e:  # noqa: BLE001
                    status = f"ошибка: {e}"
                    time.sleep(2)
            results.append((group, name, city, cnt))
            done += 1
            streak = 0 if cnt is not None else streak + 1
            shown = f"{cnt:,}".replace(",", " ") if cnt is not None else "—"
            flag = "" if status == "ок" else f"  [{status}]"
            print(f"   {group}/{name:22} {shown:>9}{flag}")
            if streak >= 12:
                print("\n⚠ 12 блоков подряд — IP Авито забанен, останавливаюсь.")
                break
            if done < total:
                time.sleep(config.REQUEST_DELAY)
    # сентинел __ABORT__ в записи не попадёт (фильтр None и спец-имени)
    return [r for r in results if r[1] != "__ABORT__"]


def write_avito(results, date_label):
    """Авито. results — (group, name, city, count). Лист config.AVITO_SHEET."""
    wb = open_or_create(config.OUTPUT_FILE, config.AVITO_SHEET)
    ws = wb[config.AVITO_SHEET]
    history, date_headers = _merge_history(ws, results, date_label)
    # «Вся Россия» — первой в каждом блоке категории, затем 16 городов
    city_order = [config.RUSSIA_LABEL] + [n for n, _, _ in config.CITIES]
    ordered, seen = [], set()
    for group, name, *_ in config.AVITO_CATEGORIES:
        for city in city_order:
            k = (group, name, city)
            if k in history and k not in seen:
                ordered.append(k)
                seen.add(k)
    for k in history:
        if k not in seen:
            ordered.append(k)
            seen.add(k)
    _save_matrix(ws, "Авито_данные", "Город", history, date_headers, ordered,
                 russia_label=config.RUSSIA_LABEL,
                 exclude_label=config.RUSSIA_LABEL)
    return _safe_save(wb, config.OUTPUT_FILE)


def _domclick_count(title):
    if "Ошибка" in title or "404" in title or "Доступ" in title:
        return None
    m = re.search(r"([\d  \s]{2,}?)\s*(?:объявлен|предложен)", title)
    if m:
        d = re.sub(r"\D", "", m.group(1))
        return int(d) if d else None
    return None


def write_domclick(results, date_label):
    """Домклик. results — (group, name, city, count). Лист config.DOMCLICK_SHEET."""
    wb = open_or_create(config.OUTPUT_FILE, config.DOMCLICK_SHEET)
    ws = wb[config.DOMCLICK_SHEET]
    history, date_headers = _merge_history(ws, results, date_label)
    # «Вся Россия» — первой в каждом блоке категории, затем 16 городов
    city_order = [config.RUSSIA_LABEL] + [n for n, _, _ in config.CITIES]
    ordered, seen = [], set()
    for group, name, *_ in config.DOMCLICK_CATEGORIES:
        for city in city_order:
            k = (group, name, city)
            if k in history and k not in seen:
                ordered.append(k)
                seen.add(k)
    for k in history:
        if k not in seen:
            ordered.append(k)
            seen.add(k)
    _save_matrix(ws, "Домклик_данные", "Город", history, date_headers, ordered,
                 russia_label=config.RUSSIA_LABEL,
                 exclude_label=config.RUSSIA_LABEL)
    return _safe_save(wb, config.OUTPUT_FILE)


def _domclick_warmup():
    """Проходит Qrator через undetected-chromedriver ОДИН раз и возвращает
    (User-Agent, dict cookies). Кука qrator_jsr ставится на .domclick.ru —
    действует на ВСЕ города-поддомены, поэтому прогрев нужен один."""
    import undetected_chromedriver as uc
    last = None
    for attempt in range(5):  # uc иногда не докачивает драйвер (ContentTooShortError)
        drv = None
        try:
            opts = uc.ChromeOptions()
            opts.add_argument("--lang=ru-RU")
            opts.add_argument("--window-size=1300,900")
            drv = uc.Chrome(options=opts, headless=False, use_subprocess=True,
                            version_main=148)
            drv.get("https://ekaterinburg.domclick.ru/")
            time.sleep(9)  # прогрев Qrator (JS-челлендж)
            ua = drv.execute_script("return navigator.userAgent")
            cookies = {c["name"]: c["value"] for c in drv.get_cookies()}
            return ua, cookies
        except Exception as e:  # noqa: BLE001
            last = e
            print(f"   uc-прогрев попытка {attempt + 1}/5 не удалась: "
                  f"{type(e).__name__}; повтор…")
            time.sleep(5)
        finally:
            if drv is not None:
                try:
                    drv.quit()
                except Exception:  # noqa: BLE001
                    pass
    raise RuntimeError(f"не удалось прогреть Домклик через uc: {last}")


def _domclick_session():
    """requests-сессия с Qrator-куками из uc-прогрева (быстрые запросы без браузера)."""
    import requests
    ua, cookies = _domclick_warmup()
    s = requests.Session()
    s.headers.update({
        "User-Agent": ua, "Accept-Language": "ru-RU,ru;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"})
    for c, v in cookies.items():
        s.cookies.set(c, v, domain=".domclick.ru")
    return s


def _domclick_get(session, url, tries=4):
    """Качает только начало страницы (до </title>) и достаёт счётчик из title.
    Ретраи при обрыве соединения / Qrator-заглушке. Возвращает int или None."""
    for _ in range(tries):
        try:
            r = session.get(url, timeout=30, stream=True)
            buf = b""
            for chunk in r.iter_content(chunk_size=4096):
                buf += chunk
                if b"</title>" in buf:
                    break
            r.close()
            text = buf.decode("utf-8", "replace")
            m = re.search(r"<title>([^<]*)", text)
            title = m.group(1) if m else ""
            cnt = _domclick_count(title)
            if cnt is not None:
                return cnt
            if (not title) or "Доступ" in title:  # Qrator/обрыв — повторить
                time.sleep(1.5)
                continue
            return None  # обычная страница без числа (0 объявлений и т.п.)
        except Exception:  # noqa: BLE001  (RemoteDisconnected и пр.)
            time.sleep(1.5)
    return None


def _collect_domclick_city(session, sub, city, cats):
    """Собирает один город. Возвращает (rows, fails)."""
    rows, fails = [], 0
    for group, name, path in cats:
        cnt = _domclick_get(session, f"https://{sub}.domclick.ru/{path}")
        rows.append((group, name, city, cnt))
        if cnt is None:
            fails += 1
        shown = f"{cnt:,}".replace(",", " ") if cnt is not None else "—"
        print(f"   {group}/{name:22} {shown:>9}{'' if cnt is not None else '  [—]'}")
        time.sleep(config.REQUEST_DELAY * 0.5)
    return rows, fails


def write_domclick_cities(results, date_label):
    """Лист «Домклик города (РФ)»: количество объявлений по каждому городу ×
    категории (то, из чего складывается «вся Россия»). Аналог листа регионов
    ЦИАН: умная таблица + строка «Итого» = всероссийская сумма (самопроверка).
    results — (group, name, city_display, count)."""
    wb = open_or_create(config.OUTPUT_FILE, config.DOMCLICK_CITIES_SHEET)
    ws = wb[config.DOMCLICK_CITIES_SHEET]
    history, date_headers = _merge_history(ws, results, date_label)
    # порядок: блоки категорий (как в DOMCLICK_CATEGORIES), внутри — города
    # в порядке DOMCLICK_RUSSIA_SUBS (по рус.названию)
    city_order = [config.DOMCLICK_SUB_NAMES.get(s, s) for s in config.DOMCLICK_RUSSIA_SUBS]
    ordered, seen = [], set()
    for group, name, *_ in config.DOMCLICK_CATEGORIES:
        for city in city_order:
            k = (group, name, city)
            if k in history and k not in seen:
                ordered.append(k)
                seen.add(k)
    for k in history:
        if k not in seen:
            ordered.append(k)
            seen.add(k)
    _save_matrix(ws, "Домклик_города", "Город", history, date_headers, ordered,
                 russia_label=None, exclude_label=None)
    return _safe_save(wb, config.OUTPUT_FILE)


def collect_domclick(date_label):
    """Домклик «по-умному»: Qrator проходим через uc ОДИН раз (берём куки),
    дальше быстрый requests с этими куками — счётчик из <title>, читаем только
    до </title>. В ~4 раза быстрее браузера. Сохранение ПОСЛЕ КАЖДОГО ГОРОДА."""
    cats = config.DOMCLICK_CATEGORIES
    total = len(config.CITIES) * len(cats)
    print(f"Домклик: {len(config.CITIES)} городов × {len(cats)} категорий = {total} "
          f"запросов (uc-прогрев + быстрый requests, ~7-9 мин). "
          f"Сохранение после каждого города.")
    print("=" * 64)
    session = _domclick_session()
    done = 0
    for city, _sub, _code in config.CITIES:
        sub = config.DOMCLICK_CITY_SUB.get(city)
        if not sub:
            continue
        print(f"\n[{city}]")
        rows, fails = _collect_domclick_city(session, sub, city, cats)
        # весь город пуст -> куки протухли: перепрогрев и один повтор
        if fails == len(cats):
            print("   ⚠ весь город пуст — обновляю Qrator-куки (перепрогрев)…")
            session = _domclick_session()
            rows, _ = _collect_domclick_city(session, sub, city, cats)
        done += len(rows)
        write_domclick(rows, date_label)
        print(f"   -> сохранено ({city})")
    return done


def collect_domclick_russia(date_label):
    """«Вся Россия» по Домклику = СУММА по всем городам-поддоменам.
    У Домклика счётчик задаётся поддоменом-городом (кука региона игнорируется,
    поддоменов уровня области нет). Листинги разных городов непересекающиеся,
    поэтому сумма по всем городам = вся Россия. Один uc-прогрев куки → быстрый
    requests. Промежуточное сохранение бегущей суммы каждые 15 городов."""
    cats = config.DOMCLICK_CATEGORIES
    subs = config.DOMCLICK_RUSSIA_SUBS
    total = len(subs) * len(cats)
    print(f"Домклик ВСЯ РОССИЯ: сумма по {len(subs)} городам × {len(cats)} категорий "
          f"= {total} запросов (~50-70 мин). Бегущая сумма сохраняется каждые 15 городов.")
    print("=" * 64)
    session = _domclick_session()
    sums = {(g, n): None for g, n, _ in cats}
    detail = {}  # (g, n, city_display) -> count (по-городовая разбивка для листа)

    def add_city(sub):
        city = config.DOMCLICK_SUB_NAMES.get(sub, sub)
        fails = 0
        for g, n, path in cats:
            cnt = _domclick_get(session, f"https://{sub}.domclick.ru/{path}")
            detail[(g, n, city)] = cnt
            if cnt is None:
                fails += 1
            else:
                sums[(g, n)] = (sums[(g, n)] or 0) + cnt
            time.sleep(config.REQUEST_DELAY * 0.4)
        return fails

    def russia_rows():
        return [(g, n, config.RUSSIA_LABEL, sums[(g, n)]) for g, n, _ in cats]

    def detail_rows():
        return [(g, n, city, c) for (g, n, city), c in detail.items()]

    def save(tag):
        # пересчитываем сумму из detail (чтобы при перепрогреве города не задвоить)
        for key in sums:
            sums[key] = None
        for (g, n, _city), c in detail.items():
            if c is not None:
                sums[(g, n)] = (sums[(g, n)] or 0) + c
        write_domclick(russia_rows(), date_label)       # итог «Вся Россия» -> лист «Домклик»
        write_domclick_cities(detail_rows(), date_label)  # по-городовая разбивка -> лист городов
        print(f"   -> сохранено ({tag})")

    for i, sub in enumerate(subs, 1):
        try:
            fails = add_city(sub)
            if fails == len(cats):  # куки протухли — перепрогрев и повтор города
                session = _domclick_session()
                add_city(sub)
        except Exception as e:  # noqa: BLE001
            print(f"   ! {sub}: {e}")
        print(f"  [{i}/{len(subs)}] {sub}")
        if i % 15 == 0:
            save(f"{i}/{len(subs)} городов")
    save(f"итог, {len(subs)} городов")
    print(f"\nДомклик «Вся Россия» + лист городов сохранены (сумма {len(subs)} городов).")
    return total


def _avito_price_call(drv):
    """Возвращает функцию call(url, body) — POST-fetch в контексте залогиненной
    страницы Авито (куки идут сами, анти-бот пройден)."""
    import json as _json
    drv.set_script_timeout(40)

    def call(url, body):
        js = ("const cb=arguments[arguments.length-1];"
              "fetch(arguments[0],{method:'POST',headers:{'content-type':'application/json'},"
              "credentials:'include',body:arguments[1]})"
              ".then(r=>r.text()).then(t=>cb(t)).catch(e=>cb('ERR:'+e));")
        return drv.execute_async_script(js, url, _json.dumps(body))
    return call


def _avito_pub_read(drv, item):
    """Один опрос публикации getAvailableBBL → кортеж (14дн,30дн,60дн) цен.
    Публикация ДЕТЕРМИНИРОВАНА локацией и обновляется МГНОВЕННО (без async-лага),
    поэтому это надёжный быстрый детектор «город реально переключился»."""
    import json as _json
    call = _avito_price_call(drv)
    try:
        bbl = _json.loads(call(f"{config.AVITO_BASE}/web/2/bbl-api/getAvailableBBL",
                               {"itemId": item}))
        by_ttl = {c.get("ttl"): c.get("priceOrigin") for c in bbl.get("configs", [])}
        return (by_ttl.get(14), by_ttl.get(30), by_ttl.get(60))
    except Exception:  # noqa: BLE001
        return None


def _avito_promo_read(drv, item):
    """Один опрос setups → список (title, price:int|None) из 3 пресетов продвижения.
    Используется для baseline ПЕРЕД сменой города (см. _avito_price_extract)."""
    import json as _json
    call = _avito_price_call(drv)

    def rub(s):
        d = re.sub(r"\D", "", str(s))
        return int(d) if d else None
    try:
        sd = _json.loads(call(f"{config.AVITO_BASE}/web/1/bbip/private/setups",
                              {"itemId": item, "vasFrom": config.AVITO_PRICE_VASFROM,
                               "s": ["bbl"]}))
        return [(p.get("title", "?"), rub(p.get("oldPriceFormatted")))
                for p in sd.get("data", {}).get("presets", [])]
    except Exception:  # noqa: BLE001
        return []


def _avito_price_extract(drv, item, promo_baseline=None):
    """Вызывает 3 private-API Авито через fetch() в контексте залогиненной
    страницы и возвращает список (Группа, Параметр, value:int|None).

    promo_baseline — значения продвижения, снятые ДО смены города (см.
    _avito_promo_read). Цена продвижения = АСИНХРОННЫЙ прогноз: после сохранения
    нового адреса сервер ещё какое-то время отдаёт СТАРОЕ (от прошлого города)
    значение, и оно стабильно — поэтому «опрос до 2 совпадений» ловил неверное
    число. С baseline опрашиваем, пока значение не ОТЛИЧИТСЯ от baseline (=сервер
    реально пересчитал под новый город) И затем стабилизируется."""
    import json as _json
    call = _avito_price_call(drv)

    def rub(s):  # "1 660 ₽" / "1 660" -> 1660
        d = re.sub(r"\D", "", str(s))
        return int(d) if d else None

    rows, vf = [], config.AVITO_PRICE_VASFROM
    # 1) публикация 14/30/60
    bbl = _json.loads(call(f"{config.AVITO_BASE}/web/2/bbl-api/getAvailableBBL",
                           {"itemId": item}))
    by_ttl = {c["ttl"]: c.get("priceOrigin") for c in bbl.get("configs", [])}
    for ttl in (14, 30, 60):
        rows.append(("Публикация", f"{ttl} дней", by_ttl.get(ttl)))
    # 2) продвижение (3 пресета за 5 дней). Прогноз отдаётся с НЕСКОЛЬКИХ кэш-узлов
    #    Авито, которые после смены города МИНУТАМИ расходятся: значение осциллирует
    #    между истинным и стале-значениями других городов. Ждём СХОДИМОСТИ узлов —
    #    NEED одинаковых чтений подряд И прошло >= MIN_T c (раньше бывает ЛОЖНАЯ
    #    сходимость на чужом устойчивом значении). Не сошлось за BUDGET → прогноз НЕ
    #    доверяем (цена=None, город на ручной добор; публикацию/выделение это не
    #    трогает). Метод сверен на эталоне: Воронеж сходится к 1060/2075/4190 за ~125 c.
    # ПАУЗА-«ОТСТАИВАНИЕ»: осцилляция/устаревание возникают из-за быстрого перебора
    #    городов подряд (частые смены засоряют кэш-узлы). Если дать прогнозу «осесть»
    #    после смены адреса — он становится стабильным и ВЕРНЫМ сразу (как при ручной
    #    правке у пользователя). Поэтому ждём REST c, потом проверяем стабильность.
    REST, NEED, BUDGET = config.AVITO_PRICE_REST, 6, 150
    time.sleep(REST)
    seen, presets, converged = [], [], False
    t0 = time.time()
    while time.time() - t0 < BUDGET:
        sd = _json.loads(call(f"{config.AVITO_BASE}/web/1/bbip/private/setups",
                              {"itemId": item, "vasFrom": vf, "s": ["bbl"]}))
        cur = [(p.get("title", "?"), rub(p.get("oldPriceFormatted")))
               for p in sd.get("data", {}).get("presets", [])]
        if cur:
            presets = cur
            seen.append(tuple(cur))
            run = 1
            for j in range(len(seen) - 2, -1, -1):
                if seen[j] == seen[-1]:
                    run += 1
                else:
                    break
            if run >= NEED:  # после отстаивания значение стабильно почти сразу
                converged = True
                break
        time.sleep(6)
    if converged:
        print(f"   ✓ продвижение сошлось за {int(time.time()-t0)} c")
        for title, price in presets:
            rows.append(("Продвижение (5 дн)", title, price))
    else:
        print(f"   ⚠ продвижение НЕ сошлось за {int(time.time()-t0)} c — на ручной добор")
        titles = [t for t, _ in presets] or ["Оценить эффект продвижения",
                  "Выбирают чаще всего", "Сильнее, чем у конкурентов"]
        for title in titles:
            rows.append(("Продвижение (5 дн)", title, None))
    # 3) выделение/XL
    cfg = _json.loads(call(f"{config.AVITO_BASE}/web/1/vas/configurator",
                           {"itemId": item, "vasFrom": vf, "s": ["bbl"],
                            "referer": f"{config.AVITO_BASE}/bbl/{item}/period"}))
    for w in cfg.get("data", {}).get("widgets", []):
        nm = w.get("name") or (w.get("data") or {}).get("name")
        op = w.get("oldPrice")
        if op is None:
            op = (w.get("data") or {}).get("oldPrice")
        if nm and op is not None:
            rows.append(("Выделение", nm, rub(op)))
    return rows


def collect_avito_price(city):
    """Цены Авито (публикация/продвижение/выделение) для текущего города
    объявления. headless undetected-chromedriver + cookie авторизации
    (config.AVITO_PRICE_COOKIE). Город задаёт ПОЛЬЗОВАТЕЛЬ в самом объявлении —
    мы лишь подписываем строки этим городом. Возвращает (Группа,Параметр,city,value)."""
    import undetected_chromedriver as uc
    try:
        raw = open(config.AVITO_PRICE_COOKIE, encoding="utf-8").read().strip()
    except FileNotFoundError:
        print(f"Нет {config.AVITO_PRICE_COOKIE} — нужны cookie авторизации Авито.")
        return []
    cookies = {k.strip(): v.strip() for k, v in
               (p.strip().split("=", 1) for p in raw.split(";") if "=" in p)}
    item = config.AVITO_PRICE_ITEM
    print(f"Avito Price [{city}], объявление {item} — headless-браузер…")

    last_err = None
    for attempt in range(3):  # uc+Chrome149 нестабилен — ретраим запуск
        drv = None
        try:
            opts = uc.ChromeOptions()
            opts.add_argument("--lang=ru-RU")
            opts.add_argument("--headless=new")
            drv = uc.Chrome(options=opts, headless=True, use_subprocess=True,
                            version_main=149)
            drv.get("https://www.avito.ru/")
            time.sleep(7)
            for k, v in cookies.items():
                try:
                    drv.add_cookie({"name": k, "value": v, "domain": ".avito.ru"})
                except Exception:  # noqa: BLE001
                    pass
            drv.get("https://www.avito.ru/")
            time.sleep(3)
            rows = _avito_price_extract(drv, item)
            ok = sum(1 for *_, v in rows if v is not None)
            if ok == 0:
                raise RuntimeError("0 цен — возможно протух sessid или firewall")
            for g, n, v in rows:
                shown = f"{v:,}".replace(",", " ") if v is not None else "—"
                print(f"   {g} / {n:32} {shown:>9}")
            return [(g, n, city, v) for g, n, v in rows]
        except Exception as e:  # noqa: BLE001
            last_err = e
            print(f"   попытка {attempt+1}/3 не удалась: {str(e)[:80]}")
            time.sleep(3)
        finally:
            if drv is not None:
                try:
                    drv.quit()
                except Exception:  # noqa: BLE001
                    pass
    print(f"Не удалось собрать цены: {last_err}")
    return []


AVITO_PRICE_COLS = (
    "Публикация 14 дн", "Публикация 30 дн", "Публикация 60 дн",
    "Продвижение: Оценить (5 дн)", "Продвижение: Выбирают чаще (5 дн)",
    "Продвижение: Сильнее (5 дн)", "Выделение цветом (7 дн)", "XL (7 дн)",
)


def write_avito_price(results, date_label):
    """Лист «Avito Price» — ШИРОКИЙ формат как CIAN Price: город по строкам,
    8 цен по столбцам + дата обновления. Снимок с накоплением по городам
    (повторный прогон города обновляет его строку). results — 8 кортежей
    (Группа, Параметр, Город, value) в каноническом порядке."""
    city = results[0][2]
    vals = [v for *_, v in results]
    cols = list(AVITO_PRICE_COLS)
    wb = open_or_create(config.OUTPUT_FILE, config.AVITO_PRICE_SHEET)
    ws = wb[config.AVITO_PRICE_SHEET]
    # накопление: читаем уже сохранённые города (если лист уже широкий)
    data = {}
    if ws.cell(1, 1).value == "Город":
        for r in range(2, ws.max_row + 1):
            c = ws.cell(r, 1).value
            if not c or "Итого" in str(c):
                continue
            data[c] = [ws.cell(r, j).value for j in range(2, 2 + len(cols) + 1)]
    new = vals + [date_label]
    old = data.get(city)
    if old:   # сохранить уже введённые значения (напр. продвижение вручную), где new=None
        new = [n if n is not None else (old[i] if i < len(old) else None)
               for i, n in enumerate(new)]
    data[city] = new
    # перезапись листа
    for tname in list(ws.tables):
        del ws.tables[tname]
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)
    titles = ["Город"] + cols + ["Обновлено"]
    for c, t in enumerate(titles, 1):
        cell = ws.cell(1, c, t)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER
        cell.alignment = Alignment(horizontal="center" if c >= 2 else "left",
                                   wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 42
    ws.column_dimensions["A"].width = 20
    for c in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.column_dimensions[get_column_letter(len(cols) + 2)].width = 12
    ws.freeze_panes = "B2"
    names = [n for n, _, _ in config.CITIES]
    order = names + [c for c in data if c not in set(names)]
    r = 2
    for c in order:
        if c not in data:
            continue
        ws.cell(r, 1, c).border = BORDER
        for j, val in enumerate(data[c], start=2):
            cell = ws.cell(r, j, val)
            cell.border = BORDER
            if isinstance(val, int):
                cell.number_format = "# ##0"
            cell.alignment = Alignment(
                horizontal="right" if j <= len(cols) + 1 else "center")
        r += 1
    last = r - 1
    tab = Table(displayName="AvitoPrice_данные",
                ref=f"A1:{get_column_letter(len(titles))}{last}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True,
                                        showColumnStripes=False)
    ws.add_table(tab)
    return _safe_save(wb, config.OUTPUT_FILE)


def _latest_col_letter(wb, sheet_name):
    """Буква последнего столбца-даты в листе-источнике (для формул-ссылок)."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    last = None
    c = 4
    while ws.cell(1, c).value is not None:
        last = c
        c += 1
    return get_column_letter(last) if last else None


def _sumifs(sheet, col, group, name, city_ref):
    """Одна формула SUMIFS к листу-источнику по Группа+Категория+Город."""
    g = group.replace('"', '""')
    n = name.replace('"', '""')
    return (f"SUMIFS('{sheet}'!${col}:${col},'{sheet}'!$A:$A,\"{g}\","
            f"'{sheet}'!$B:$B,\"{n}\",'{sheet}'!$C:$C,{city_ref})")


def build_compare(date_label):
    """Сводный лист «Сравнение»: общие категории × города × 4 источника.
    Значения — ФОРМУЛЫ-ссылки (SUMIFS) на листы-источники (последний столбец-дата),
    «Лидер» — формула, максимум в строке подсвечивается условным форматированием."""
    from openpyxl.formatting.rule import FormulaRule
    wb = open_or_create(config.OUTPUT_FILE, config.COMPARE_SHEET)
    cols = {s: _latest_col_letter(wb, s) for s in config.COMPARE_SOURCES}
    # у каких источников есть строка «Вся Россия» (Домклика нет → покажем «—»)
    src_russia = {}
    for s in config.COMPARE_SOURCES:
        has = False
        if s in wb.sheetnames:
            for r in wb[s].iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
                if r[0] == config.RUSSIA_LABEL:
                    has = True
                    break
        src_russia[s] = has
    ws = wb[config.COMPARE_SHEET]
    for tname in list(ws.tables):
        del ws.tables[tname]
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    titles = ["Сделка", "Категория", "Город"] + config.COMPARE_SOURCES + ["Лидер"]
    for col, t in enumerate(titles, 1):
        c = ws.cell(1, col, t)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(horizontal="center" if col >= 4 else "left")
    ws.freeze_panes = "D2"
    for col, w in zip(range(1, len(titles) + 1), [12, 26, 20, 11, 11, 11, 11, 12]):
        ws.column_dimensions[get_column_letter(col)].width = w

    nsrc = len(config.COMPARE_SOURCES)
    first_src, last_src = get_column_letter(4), get_column_letter(3 + nsrc)
    leader_col = 4 + nsrc
    # «Вся Россия» — первой строкой в каждом блоке категории, затем 16 городов
    city_order = [config.RUSSIA_LABEL] + [c[0] for c in config.CITIES]
    row, prev_block = 2, None
    for deal, name, mp in config.COMPARE_ROWS:
        for city in city_order:
            is_russia = city == config.RUSSIA_LABEL
            ws.cell(row, 1, deal)
            ws.cell(row, 2, name)
            ws.cell(row, 3, city)
            city_ref = f"$C{row}"
            for j, s in enumerate(config.COMPARE_SOURCES, start=4):
                mapping = mp.get(s)
                cell = ws.cell(row, j)
                if mapping is None or cols[s] is None or (is_russia and not src_russia[s]):
                    cell.value = "—"
                else:
                    keys = [mapping] if isinstance(mapping, tuple) else mapping
                    parts = [_sumifs(s, cols[s], g, n, city_ref) for g, n in keys]
                    cell.value = "=" + "+".join(parts)
                    cell.number_format = "# ##0"
                cell.alignment = Alignment(horizontal="right")
            # Лидер — формула: имя источника с максимумом в строке
            rng = f"{first_src}{row}:{last_src}{row}"
            ws.cell(row, leader_col,
                    f"=IFERROR(INDEX(${first_src}$1:${last_src}$1,"
                    f"MATCH(MAX({rng}),{rng},0)),\"\")")
            block = (deal, name)
            new_block = block != prev_block
            for col in range(1, len(titles) + 1):
                top = TOPSEP if new_block else THIN
                ws.cell(row, col).border = Border(left=THIN, right=THIN, top=top, bottom=THIN)
            if new_block:
                ws.cell(row, 1).font = GROUP_FONT
                ws.cell(row, 2).font = CAT_FONT
                for col in range(1, 4):
                    ws.cell(row, col).fill = BLOCK_FILL
                prev_block = block
            row += 1

    last = row - 1
    # подсветка максимума в строке (зелёный жирный) — условным форматированием
    rule = FormulaRule(
        formula=[f"AND(ISNUMBER({first_src}2),{first_src}2=MAX($"
                 f"{first_src}2:${last_src}2))"],
        font=Font(bold=True, color="1F7A1F"))
    ws.conditional_formatting.add(f"{first_src}2:{last_src}{last}", rule)

    tab = Table(displayName="Сравнение_данные",
                ref=f"A1:{get_column_letter(len(titles))}{last}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=False,
                                        showColumnStripes=False)
    ws.add_table(tab)
    wb.move_sheet(config.COMPARE_SHEET, -(wb.sheetnames.index(config.COMPARE_SHEET)))
    return _safe_save(wb, config.OUTPUT_FILE)


def _read_cookie():
    try:
        return open(config.COOKIE_FILE, encoding="utf-8").read().strip()
    except FileNotFoundError:
        return None


def collect_prices():
    """Цены публикации (тарифы) по 16 городам × сделка × тип недвижимости.
    Требует cookie авторизованной сессии (файл config.COOKIE_FILE)."""
    import json
    import requests
    cookie = _read_cookie()
    if not cookie:
        print(f"Нет файла {config.COOKIE_FILE} — цены требуют авторизованную cookie.")
        return []
    session = requests.Session()
    session.headers.update({
        "User-Agent": UA, "Content-Type": "application/json",
        "Accept": "application/json", "Referer": "https://my.cian.ru/price",
        "Origin": "https://my.cian.ru", "Cookie": cookie})

    total = len(config.CITIES) * len(config.PRICE_DEALS) * len(config.PRICE_OFFERS)
    print(f"Цены: {len(config.CITIES)} городов × {len(config.PRICE_DEALS)} сделок × "
          f"{len(config.PRICE_OFFERS)} типов = {total} запросов")
    print("=" * 64)

    rows, done = [], 0
    for city, _sub, _code in config.CITIES:
        polygon = config.PRICE_POLYGONS.get(city)
        if polygon is None:
            continue
        for deal, deal_label in config.PRICE_DEALS:
            for offer, offer_label in config.PRICE_OFFERS:
                done += 1
                payload = {"dealType": deal, "offerType": offer, "polygonId": polygon}
                data = None
                for _ in range(4):
                    try:
                        r = session.post(config.PRICE_API,
                                         data=json.dumps(payload), timeout=30)
                        if r.status_code == 400 and "не определен" in r.text:
                            print("\n⚠ Cookie недействителен/протух "
                                  "(«Пользователь не определён»).")
                            print(f"  Обнови {config.COOKIE_FILE} "
                                  f"(см. комментарий в config.py).")
                            return []
                        data = r.json()
                        break
                    except Exception:  # noqa: BLE001
                        time.sleep(1.0)
                pl = (data or {}).get("priceList") or {}
                head = [h.get("text") for h in pl.get("head", [])]
                tariffs = head[1:]
                note = (data or {}).get("note")
                body = pl.get("body", [])
                for body_row in body:
                    cat = body_row[0].get("text")
                    vals = {}
                    for idx, cell in enumerate(body_row[1:]):
                        val = cell.get("price")
                        if val is None:
                            val = cell.get("from")
                        if idx < len(tariffs):
                            vals[tariffs[idx]] = val
                    rows.append({"deal": deal_label, "offer": offer_label,
                                 "category": cat, "city": city,
                                 "vals": vals, "note": note})
                print(f"[{done}/{total}] {city} · {deal_label} · {offer_label}: "
                      f"{len(body)} категорий")
                time.sleep(0.15)
    return rows


def _yandex_count(html, mode):
    """Число объявлений со страницы Яндекс.Недвижимости."""
    if mode == "newbuilding":
        m = re.search(r'"newbuildingOffersCount":(\d+)', html)
        return int(m.group(1)) if m else None
    m = re.search(r"<title>([^<]*)</title>", html)
    title = m.group(1) if m else ""
    if "Ошибка" in title or "404" in title:
        return None
    # «— 69 320 объявлений…» -> число; «— объявления домов…» (нет числа) -> 0
    m2 = re.search(r"—\s*([\d\s  ]*?)\s*"
                   r"(?:объявлен|вариант|предложен|домов|гараж|квартир|"
                   r"комнат|участк|объект)", title)
    if m2:
        d = re.sub(r"\D", "", m2.group(1))
        return int(d) if d else 0
    return None


def _yandex_one(url, mode):
    """Один запрос к Яндексу через curl --compressed с ретраями. -> count|None.
    Отсеивает капчу и «облегчённые» анти-бот страницы (без <title>)."""
    for _ in range(3):
        try:
            html, _final = fetch_html(url)
            low = html.lower()
            if "showcaptcha" in low or "<title>" not in low:
                time.sleep(2)
                continue
            return _yandex_count(html, mode)
        except Exception:  # noqa: BLE001
            time.sleep(1)
    return None


def collect_yandex_regions():
    """Разбивка по всем 85 субъектам РФ. Возвращает (region_rows, russia_rows):
    region_rows = (group, name, регион, count) для листа регионов;
    russia_rows = (group, name, «Вся Россия», сумма) для главного листа Яндекса."""
    regions = list(config.YANDEX_REGION_SLUG.items())
    cats = config.YANDEX_CATEGORIES
    total = len(regions) * len(cats)
    print(f"Яндекс РФ: {len(regions)} субъектов × {len(cats)} категорий "
          f"= {total} запросов (~35-40 мин)")
    print("=" * 64)
    region_rows, sums, done = [], {}, 0
    for rname, slug in regions:
        for group, name, catslug, mode in cats:
            url = f"{config.YANDEX_BASE}/{slug}/{catslug}/"
            cnt = _yandex_one(url, mode)
            region_rows.append((group, name, rname, cnt))
            if cnt is not None:
                sums[(group, name)] = sums.get((group, name), 0) + cnt
            done += 1
            time.sleep(0.15)
        got = sum(1 for r in region_rows[-len(cats):] if r[3] is not None)
        print(f"  [{done // len(cats):>2}/{len(regions)}] {rname:24} ({got}/{len(cats)})")
    russia_rows = [(g, n, config.RUSSIA_LABEL, s) for (g, n), s in sums.items()]
    return region_rows, russia_rows


def collect_yandex():
    """Количество объявлений Яндекс.Недвижимости по 16 городам × категориям.
    Через curl --compressed (gzip даёт скорость; TLS-отпечаток curl Яндекс не
    душит, в отличие от requests). Число из <title>. (group, name, city, count)."""
    total = len(config.CITIES) * len(config.YANDEX_CATEGORIES)
    print(f"Яндекс: {len(config.CITIES)} городов × {len(config.YANDEX_CATEGORIES)} "
          f"категорий = {total} запросов")
    print("=" * 64)
    results, done = [], 0
    for city, _sub, _code in config.CITIES:
        slug = config.YANDEX_CITY_SLUG.get(city)
        if not slug:
            continue
        print(f"\n[{city}]")
        for group, name, catslug, mode in config.YANDEX_CATEGORIES:
            url = f"{config.YANDEX_BASE}/{slug}/{catslug}/"
            cnt, status = None, "ошибка"
            for _ in range(3):
                try:
                    html, _final = fetch_html(url)
                    low = html.lower()
                    if "showcaptcha" in low:
                        status = "КАПЧА"
                        time.sleep(2)
                        continue
                    if "<title>" not in low:  # «облегчённая» страница анти-бота
                        status = "урезано (анти-бот)"
                        time.sleep(2)
                        continue
                    cnt = _yandex_count(html, mode)
                    status = "ок" if cnt is not None else "нет числа"
                    break
                except Exception as e:  # noqa: BLE001
                    status = f"ошибка: {e}"
                    time.sleep(1)
            results.append((group, name, city, cnt))
            done += 1
            shown = f"{cnt:,}".replace(",", " ") if cnt is not None else "—"
            flag = "" if status == "ок" else f"  [{status}]"
            print(f"   {group}/{name:24} {shown:>9}{flag}")
            if done < total:
                time.sleep(0.2)
    return results


def collect_cities():
    """Парсит все категории по 16 городам (через curl). Возвращает список
    (group, name, city, count_or_None)."""
    total_tasks = len(config.CATEGORIES) * len(config.CITIES)
    print(f"Города: {len(config.CITIES)} × {len(config.CATEGORIES)} категорий "
          f"= {total_tasks} запросов")
    print("=" * 64)
    results, done = [], 0
    for group, name, kind, value in config.CATEGORIES:
        print(f"\n[{group}] {name}")
        for city, sub, code in config.CITIES:
            url = config.build_url(kind, value, sub, code)
            count, status = parse_one(url)
            results.append((group, name, city, count))
            done += 1
            shown = f"{count:,}".replace(",", " ") if count is not None else "—"
            flag = "" if status == "ок" else f"  [{status}]"
            print(f"   {city:<18} {shown:>9}{flag}")
            if done < total_tasks:
                time.sleep(config.REQUEST_DELAY)
    return results


def main():
    # режим: cities (по умолчанию) | russia | all | regions
    mode = sys.argv[1].lower() if len(sys.argv) > 1 else "cities"
    valid = ("cities", "russia", "all", "regions", "prices", "yandex",
             "yandex_rf", "avito", "avito_price", "domclick", "domclick_rf", "compare")
    if mode not in valid:
        print(f"Неизвестный режим: {mode}. Используйте: " + " | ".join(valid))
        sys.exit(1)

    date_label = datetime.date.today().isoformat()
    print(f"ЦИАН парсер — режим: {mode}, дата: {date_label}")
    print("=" * 64)

    # regions — отдельный лист с разбивкой по всем субъектам РФ
    if mode == "regions":
        results = collect_regions()
        ok = sum(1 for *_, c in results if c is not None)
        print("\n" + "=" * 64)
        print(f"Получено {ok} из {len(results)} значений.")
        path = write_regions(results, date_label)
        print(f"Сохранено в: {path} (лист «{config.REGIONS_SHEET}»)")
        return

    # yandex — отдельный лист с количеством объявлений Яндекс.Недвижимости
    if mode == "yandex":
        results = collect_yandex()
        ok = sum(1 for *_, c in results if c is not None)
        print("\n" + "=" * 64)
        print(f"Получено {ok} из {len(results)} значений.")
        path = write_yandex(results, date_label)
        print(f"Сохранено в: {path} (лист «{config.YANDEX_SHEET}»)")
        return

    # yandex_rf — «вся Россия» по Яндексу: лист регионов + строки «Вся Россия»
    if mode == "yandex_rf":
        region_rows, russia_rows = collect_yandex_regions()
        ok = sum(1 for *_, c in region_rows if c is not None)
        print("\n" + "=" * 64)
        print(f"Получено {ok} из {len(region_rows)} значений по регионам.")
        write_yandex_regions(region_rows, date_label)
        path = write_yandex(russia_rows, date_label)
        print(f"Сохранено: лист «{config.YANDEX_REGIONS_SHEET}» + строки «Вся Россия» "
              f"на листе «{config.YANDEX_SHEET}» -> {path}")
        return

    # compare — сводный лист сравнения 4 источников (читает существующие листы)
    if mode == "compare":
        path = build_compare(date_label)
        print(f"Сводный лист «{config.COMPARE_SHEET}» построен -> {path}")
        return

    # domclick — отдельный лист (uc-прогрев куки + быстрый requests)
    if mode == "domclick":
        n = collect_domclick(date_label)
        print("\n" + "=" * 64)
        print(f"Домклик: обработано {n} запросов. Лист «{config.DOMCLICK_SHEET}».")
        return

    # domclick_rf — строка «Вся Россия» = сумма по всем городам-поддоменам
    if mode == "domclick_rf":
        n = collect_domclick_russia(date_label)
        print("\n" + "=" * 64)
        print(f"Домклик «Вся Россия»: обработано {n} запросов. Лист «{config.DOMCLICK_SHEET}».")
        return

    # avito — отдельный лист с количеством объявлений Авито
    # доп.аргумент: список городов через запятую (дозабор), напр.
    #   python cian_parser.py avito "Самара,Омск,Пермь"
    if mode == "avito":
        only = sys.argv[2].split(",") if len(sys.argv) > 2 else None
        results = collect_avito([c.strip() for c in only] if only else None)
        ok = sum(1 for *_, c in results if c is not None)
        print("\n" + "=" * 64)
        print(f"Получено {ok} из {len(results)} значений.")
        path = write_avito(results, date_label)
        print(f"Сохранено в: {path} (лист «{config.AVITO_SHEET}»)")
        return

    # avito_price — цены Авито для города, в который СЕЙЧАС выставлено объявление.
    #   python cian_parser.py avito_price "Москва"
    if mode == "avito_price":
        if len(sys.argv) < 3:
            print('Укажите город: python cian_parser.py avito_price "Москва"')
            return
        city = sys.argv[2].strip()
        results = collect_avito_price(city)
        if results:
            path = write_avito_price(results, date_label)
            print("\n" + "=" * 64)
            print(f"Сохранено: {len([r for r in results if r[3] is not None])} цен "
                  f"для «{city}» -> {path} (лист «{config.AVITO_PRICE_SHEET}»)")
        return

    # prices — отдельный лист с ценами публикации (нужна cookie)
    if mode == "prices":
        rows = collect_prices()
        if not rows:
            print("Цены не собраны (нет/протух cookie).")
            return
        print("\n" + "=" * 64)
        print(f"Собрано {len(rows)} строк цен.")
        path = write_prices(rows, date_label)
        print(f"Сохранено в: {path} (лист «{config.PRICE_SHEET}»)")
        return

    results = []
    if mode in ("cities", "all"):
        results += collect_cities()
    if mode in ("russia", "all"):
        results += collect_russia()

    ok = sum(1 for *_, c in results if c is not None)
    print("\n" + "=" * 64)
    print(f"Получено {ok} из {len(results)} значений.")
    path = write_results(results, date_label)
    print(f"Сохранено в: {path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:  # noqa: BLE001
        print("ОШИБКА:", e)
        sys.exit(1)
